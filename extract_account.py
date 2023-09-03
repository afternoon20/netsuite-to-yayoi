import util
import openpyxl
from classes.Yayoi import Yayoi as Yayoi

# python extract_account.py
wb_netsuite = util.netsuite.load_financial_statements()
ws_netsuite_bs = None
ws_netsuite_pl = None
wb_yayoi = util.yayoi.load_financial_statements()
ws_yayoi_bs = None
ws_yayoi_pl = None
try:
    ws_netsuite_bs = wb_netsuite['BS']
    ws_netsuite_pl = wb_netsuite['PL']
    ws_yayoi_bs = wb_yayoi['BS']
    ws_yayoi_pl = wb_yayoi['PL']
except:
    print('BSまたはPLシートが存在しません。')

netsuite_account_names = util.netsuite.get_account_name(ws_netsuite_bs)
netsuite_account_names += util.netsuite.get_account_name(ws_netsuite_pl)
yayoi_account_names = util.yayoi.get_account_name(ws_yayoi_bs)
yayoi_account_names += util.yayoi.get_account_name(ws_yayoi_pl)

extract_account_workbook = openpyxl.Workbook()
extract_account_worksheet = extract_account_workbook.worksheets[0]
extract_account_worksheet.cell(1, 1).value = '登録されている勘定科目'
extract_account_worksheet.cell(2, 1).value = '弥生会計'
extract_account_worksheet.cell(2, 3).value = 'NetSuite'
extract_account_worksheet.cell(2, 4).value = '弥生会計に登録済み'

for i in range(len(yayoi_account_names)):
    extract_account_worksheet.cell(i + 3, 1).value = yayoi_account_names[i]

for i in range(len(netsuite_account_names)):
    extract_account_worksheet.cell(i + 3, 3).value = Yayoi.trim_for_yayoi(netsuite_account_names[i])
    if Yayoi.trim_for_yayoi(netsuite_account_names[i]) in yayoi_account_names:
        extract_account_worksheet.cell(i + 3, 4).value = '◯'

extract_account_workbook.save('./files/generate/extract_account.xlsx')