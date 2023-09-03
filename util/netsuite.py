import openpyxl

def load_general_ledger_work_book():
    return openpyxl.load_workbook('files/netsuite/export.xlsx')

def load_financial_statements():
    return openpyxl.load_workbook('files/netsuite/financial_statements.xlsx')

def get_start_row(ws, header_name):
    start_row = 0
    for row in ws.rows:
        for cell in row:
            if cell.value == header_name:
                start_row = cell.row
                break

    return start_row

def set_header_row(ws, header, start_row):
    global header_row
    for row in ws.rows:
        for cell in row:
            if cell.row == start_row:
                header_row = row
                break
    for cell in header_row:
        if (cell.value in header.keys()):
            header[cell.value] = cell.column

def set_account(ws, account_column = 0, start_row = 1):

    account = None
    for row in ws.iter_cols(account_column, account_column):
        for cell in row:
            if (cell.row < start_row): continue
            if (cell.value):
                account = cell.value
            else:
                cell.value = account

def get_account_name(ws):
    account_names = []
    account_row = int(1)
    for row in ws.iter_cols(account_row, account_row):
        for cell in row:
            if cell.value in ['', None, False, 0]: continue
            # TODO: 勘定科目の条件が会社によって違うか確認
            if cell.font.b == False:
                account_names.append(cell.value)

    return account_names