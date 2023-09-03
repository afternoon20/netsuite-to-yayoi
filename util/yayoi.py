import openpyxl
from classes.Yayoi import Yayoi as Yayoi
import config

def load_financial_statements():
    return openpyxl.load_workbook('files/yayoi_kaikei/financial_statements.xlsx')

def generate_yayois(transactions):
    yayois = []
    tmp_yayois = []
    for transaction in transactions:
        yayoi = Yayoi()
        yayoi.set_data(transaction)
        tmp_yayois.append(yayoi)
    yayois = arrange_by_slip_number(tmp_yayois)

    return yayois

def arrange_by_slip_number(tmp_yayois):
    arrange_yayois = {}
    for yayoi in tmp_yayois:
        if (yayoi.slip_number not in arrange_yayois):
                arrange_yayois[yayoi.slip_number] = []
        arrange_yayois[yayoi.slip_number].append(yayoi)

    return arrange_yayois


def create_yayoi_import_file(yayois):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    slip_row = int(1)
    # TODO: 事前にセル位置定義してforする
    for slip_number, yayoi_list in yayois.items():
        for yayoi in yayoi_list:
            ws.cell(slip_row, 1).value = yayoi.id_flag
            ws.cell(slip_row, 2).value = yayoi.slip_number
            ws.cell(slip_row, 3).value = yayoi.settlement_slip
            ws.cell(slip_row, 4).value = yayoi.slip_day
            ws.cell(slip_row, 5).value = yayoi.debit_name
            ws.cell(slip_row, 6).value = yayoi.debit_sub_name
            ws.cell(slip_row, 7).value = yayoi.debit_description
            ws.cell(slip_row, 8).value = yayoi.debit_tax_type
            ws.cell(slip_row, 9).value = yayoi.debit_amount
            ws.cell(slip_row, 10).value = yayoi.debit_tax
            ws.cell(slip_row, 11).value = yayoi.credit_name
            ws.cell(slip_row, 12).value = yayoi.credit_sub_name
            ws.cell(slip_row, 13).value = yayoi.credit_description
            ws.cell(slip_row, 14).value = yayoi.credit_tax_type
            ws.cell(slip_row, 15).value = yayoi.credit_amount
            ws.cell(slip_row, 16).value = yayoi.credit_tax
            ws.cell(slip_row, 17).value = yayoi.summary
            ws.cell(slip_row, 18).value = yayoi.bill_number
            ws.cell(slip_row, 19).value = yayoi.due_date
            ws.cell(slip_row, 20).value = yayoi.slip_type
            ws.cell(slip_row, 21).value = yayoi.system_source
            ws.cell(slip_row, 22).value = yayoi.memo
            ws.cell(slip_row, 23).value = yayoi.tag1
            ws.cell(slip_row, 24).value = yayoi.tag2
            ws.cell(slip_row, 25).value = yayoi.adjustment
            slip_row += 1
    wb.save('./files/generate/import_data.xlsx')

def get_account_name(ws):
    account_names = []
    account_row = int(1)
    is_list = False
    for row in ws.iter_cols(account_row, account_row):
        for cell in row:
            if (cell.value in ['', None, False, 0]) or (cell.value in config.staticlist.YAYOI_KAIKEI_FINANCIAL_STATEMENT_NON_ACCOUNT_LIST): continue
            if cell.value == '勘定科目':
                is_list = True
                continue
            if is_list:
                account_names.append(cell.value)

    return account_names